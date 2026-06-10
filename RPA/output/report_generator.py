"""
Gerador de artefatos de saída do SpaceIntel RPA.

Produz automaticamente:
  1. Relatório Excel (.xlsx) — múltiplas abas + gráfico matplotlib embutido
  2. Relatório de análise (.md) — síntese narrativa gerada pela IA
"""

import io
from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR, EXCEL_FILENAME, REPORT_FILENAME
from output.logger import get_logger

logger = get_logger("report_generator")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Paleta 2026 ───────────────────────────────────────────────────────────────
_C_DARK    = "0D1B2A"
_C_MID     = "1B3A5C"
_C_ALT     = "EAF4FB"
_C_NEG     = "C0392B"
_C_POS     = "1A7A4A"

_THIN      = Side(style="thin",   color="C5D5E8")
_MED       = Side(style="medium", color=_C_MID)
_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_HD = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)
_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(color="000000", bold=False, size=10):
    return Font(name="Arial", color=color, bold=bold, size=size)


def _style_sheet(ws, title: str = "", accent: str = _C_DARK) -> None:
    if title:
        ws.insert_rows(1); ws.insert_rows(1)
        end_col = get_column_letter(max(ws.max_column or 1, 6))
        ws.merge_cells(f"A1:{end_col}2")
        c = ws["A1"]
        c.value = title
        c.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        c.fill  = _fill(accent)
        c.alignment = _CENTER
        hdr_row = 3
    else:
        hdr_row = 1

    for cell in ws[hdr_row]:
        cell.fill      = _fill(accent)
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.alignment = _CENTER
        cell.border    = _BORDER_HD

    for i, row in enumerate(ws.iter_rows(min_row=hdr_row + 1), 1):
        bg = _fill(_C_ALT) if i % 2 == 0 else _fill("FFFFFF")
        for cell in row:
            cell.fill      = bg
            cell.font      = _font()
            cell.alignment = _LEFT
            cell.border    = _BORDER

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 5, 55)

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    ws.sheet_view.showGridLines = False


def _colorize_variation(ws, col_letter: str, start_row: int, end_row: int):
    for r in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{r}"]
        try:
            val = float(str(cell.value).replace("%","").replace("+","").replace("▼","").replace("▲","").strip())
            if val < 0:
                cell.font  = Font(name="Arial", size=10, bold=True, color=_C_NEG)
                cell.value = f"▼ {val:+.2f}%"
            elif val > 0:
                cell.font  = Font(name="Arial", size=10, bold=True, color=_C_POS)
                cell.value = f"▲ {val:+.2f}%"
        except (ValueError, TypeError):
            pass


def _build_chart_image(market_df: pd.DataFrame) -> bytes:
    """Gera gráfico de barras moderno com matplotlib → PNG em memória."""
    ativos = market_df["ativo"].tolist()
    valores = [float(v) for v in market_df["variacao_pct"].tolist()]

    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor("#0D1B2A")
    ax.set_facecolor("#0D1B2A")

    colors = ["#C0392B" if v < 0 else "#1A7A4A" for v in valores]
    bars = ax.bar(ativos, valores, color=colors, width=0.5,
                  zorder=3, edgecolor="none")

    # Data labels
    for bar, val in zip(bars, valores):
        ypos = val - 0.15 if val < 0 else val + 0.08
        ax.text(bar.get_x() + bar.get_width() / 2, ypos,
                f"{val:+.2f}%", ha="center", va="top" if val < 0 else "bottom",
                color="white", fontsize=11, fontweight="bold")

    # Linha zero
    ax.axhline(0, color="#00C2FF", linewidth=1.2, zorder=2)

    # Grid
    ax.yaxis.grid(True, color="#1B3A5C", linewidth=0.8, linestyle="--", zorder=1)
    ax.set_axisbelow(True)

    # Eixos
    ax.tick_params(colors="white", labelsize=11)
    ax.set_ylabel("Variação (%)", color="#00C2FF", fontsize=11, labelpad=10)
    ax.set_title("Variação % de Insumos Aeroespaciais",
                 color="white", fontsize=14, fontweight="bold", pad=16)

    for spine in ax.spines.values():
        spine.set_edgecolor("#1B3A5C")

    ax.yaxis.set_tick_params(colors="white")
    ax.xaxis.set_tick_params(colors="white")
    plt.setp(ax.get_xticklabels(), color="white", fontsize=11)
    plt.setp(ax.get_yticklabels(), color="#A0B4C8", fontsize=10)

    plt.tight_layout(pad=1.5)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _add_market_chart(wb, market_df: pd.DataFrame) -> None:
    if market_df.empty:
        return

    ws = wb.create_sheet("📊 Gráfico de Mercado")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00C2FF"

    # Título
    ws["B2"].value = "Variação % de Insumos Aeroespaciais"
    ws["B2"].font  = Font(name="Arial", bold=True, size=14, color="0D1B2A")
    ws.row_dimensions[2].height = 28

    # Dados de referência (coluna A/B, ocultos visualmente)
    ws["A5"].value = "Ativo"; ws["B5"].value = "Var (%)"
    ws["A5"].font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws["A5"].fill  = _fill(_C_DARK)
    ws["B5"].font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws["B5"].fill  = _fill(_C_DARK)
    for i, (_, row) in enumerate(market_df.iterrows(), start=6):
        ws[f"A{i}"].value = row["ativo"]
        ws[f"B{i}"].value = float(row["variacao_pct"])
        ws[f"A{i}"].font  = _font()
        ws[f"B{i}"].font  = _font()
        bg = _fill("FDECEA") if float(row["variacao_pct"]) < 0 else _fill("E8F5E9")
        ws[f"A{i}"].fill  = bg; ws[f"B{i}"].fill = bg

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    # Imagem matplotlib
    png_bytes = _build_chart_image(market_df)
    img = XLImage(io.BytesIO(png_bytes))
    img.anchor = "D2"
    ws.add_image(img)

    logger.info("[green]✓ Gráfico de mercado adicionado[/green]")


def generate_excel(
    launches_df: pd.DataFrame,
    market_df: pd.DataFrame,
    insights: dict[str, Any],
) -> Path:
    path = OUTPUT_DIR / EXCEL_FILENAME

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        meta = pd.DataFrame([{
            "Execução":              datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Lançamentos Coletados": len(launches_df),
            "Ativos de Mercado":     len(market_df),
            "Fonte de IA":           insights.get("fonte_analise", "N/D"),
            "Atividade Espacial":    insights.get("indice_atividade_espacial", "N/D"),
            "Sentimento de Mercado": insights.get("sentimento_mercado", "N/D"),
        }])
        meta.to_excel(writer, sheet_name="🏠 Dashboard", index=False)

        df_l = launches_df if not launches_df.empty else pd.DataFrame({"aviso": ["Nenhum lançamento coletado"]})
        df_l.to_excel(writer, sheet_name="🛸 Lançamentos", index=False)

        if not market_df.empty:
            market_df.to_excel(writer, sheet_name="📈 Mercado", index=False)

        ai_rows = [
            {"Campo": k.replace("_"," ").title(),
             "Valor": ", ".join(v) if isinstance(v, list) else str(v)}
            for k, v in insights.items()
        ]
        pd.DataFrame(ai_rows).to_excel(writer, sheet_name="🧠 Análise IA", index=False)

    wb = load_workbook(path)
    _style_sheet(wb["🏠 Dashboard"],   title="SpaceIntel RPA — Dashboard de Execução",  accent=_C_DARK)
    _style_sheet(wb["🛸 Lançamentos"], title="Missões Espaciais Coletadas",              accent=_C_MID)
    _style_sheet(wb["🧠 Análise IA"],  title="Análise Inteligente — GPT-4o",             accent="1A4A7A")

    if "📈 Mercado" in wb.sheetnames:
        ws_m = wb["📈 Mercado"]
        _style_sheet(ws_m, title="Insumos Aeroespaciais — Cotações", accent="1A5C3A")
        for col_idx, cell in enumerate(ws_m[3], start=1):
            if cell.value and "varia" in str(cell.value).lower():
                _colorize_variation(ws_m, get_column_letter(col_idx), 4, ws_m.max_row)
                break

    _add_market_chart(wb, market_df)

    # Ordenar abas
    order = ["🏠 Dashboard","🛸 Lançamentos","📈 Mercado","🧠 Análise IA","📊 Gráfico de Mercado"]
    ordered = [s for s in order if s in wb.sheetnames]
    wb._sheets.sort(key=lambda s: ordered.index(s.title) if s.title in ordered else 99)

    # Tab colors
    tab_colors = {
        "🏠 Dashboard":          "0D1B2A",
        "🛸 Lançamentos":        "1B3A5C",
        "📈 Mercado":            "1A5C3A",
        "🧠 Análise IA":         "1A4A7A",
        "📊 Gráfico de Mercado": "00C2FF",
    }
    for name, color in tab_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    wb.save(path)
    logger.info(f"[green]✓ Excel gerado: {path}[/green]")
    return path


def generate_markdown_report(
    launches_df: pd.DataFrame,
    market_df: pd.DataFrame,
    insights: dict[str, Any],
) -> Path:
    path  = OUTPUT_DIR / REPORT_FILENAME
    now   = datetime.now().strftime("%d/%m/%Y às %H:%M")
    fonte = insights.get("fonte_analise", "N/D")
    atv   = insights.get("indice_atividade_espacial", "N/D")
    sent  = insights.get("sentimento_mercado", "N/D")

    s_icon = {"Otimista":"🟢","Neutro":"🟡","Pessimista":"🔴"}.get(sent,"⚪")
    a_icon = {"Alto":"🔥","Médio":"📡","Baixo":"🌑"}.get(atv,"📡")

    lines = [
        "# 🚀 SpaceIntel RPA — Relatório de Análise Automatizada",
        "",
        "> *Gerado automaticamente pelo pipeline SpaceIntel RPA*",
        "",
        "---",
        "## 📌 Resumo da Execução",
        "",
        "| Parâmetro | Valor |",
        "|---|---|",
        f"| 🕐 **Gerado em** | {now} |",
        f"| 🤖 **Fonte de IA** | {fonte} |",
        f"| 🛸 **Lançamentos coletados** | {len(launches_df)} missões |",
        f"| 📦 **Ativos monitorados** | {len(market_df)} insumos |",
        f"| 📡 **Atividade Espacial** | {a_icon} **{atv}** |",
        f"| 📊 **Sentimento de Mercado** | {s_icon} **{sent}** |",
        "",
        "---",
        "## 🧠 Resumo Executivo",
        "",
        f"> {insights.get('resumo_executivo','N/D')}",
        "",
        "---",
        "## 🛸 Eventos Críticos Identificados",
        "",
    ]
    for item in insights.get("eventos_criticos", []):
        lines.append(f"- 🔸 {item}")

    lines += ["", "---", "## 📈 Mercado de Insumos Aeroespaciais", "",
              insights.get("analise_mercado","N/D"), ""]

    if not market_df.empty:
        lines += ["| Ativo | Preço (USD) | Variação % | |",
                  "|:---|---:|:---:|:---:|"]
        for _, row in market_df.iterrows():
            v = float(row["variacao_pct"])
            arrow = "🔴 ▼" if v < 0 else "🟢 ▲"
            lines.append(f"| **{row['ativo']}** | `${float(row['preco_atual']):,.2f}` | `{v:+.2f}%` | {arrow} |")

    lines += ["", "---", "## ⚠️ Riscos Identificados", ""]
    for r in insights.get("riscos_identificados", []):
        lines.append(f"- ⚡ {r}")

    lines += ["", "---", "## 💡 Oportunidades", ""]
    for o in insights.get("oportunidades", []):
        lines.append(f"- ✅ {o}")

    lines += ["", "---", "## ✅ Recomendações", ""]
    for i, rec in enumerate(insights.get("recomendacoes", []), 1):
        lines.append(f"{i}. {rec}")

    top = min(len(launches_df), 20)
    lines += ["", "---", f"## 🛰️ Próximos Lançamentos *(top {top} de {len(launches_df)})*", ""]
    if not launches_df.empty:
        lines += ["| # | Data | Missão | Veículo | Local |",
                  "|:---:|:---|:---|:---|:---|"]
        for i, (_, row) in enumerate(launches_df.head(20).iterrows(), 1):
            lines.append(f"| {i} | {row['data']} | **{row['missao']}** | {row['veiculo']} | {row['local']} |")

    lines += ["", "---", "<sub>",
              f"📡 Dados: Launch Library 2 API + Yahoo Finance &nbsp;|&nbsp; 🤖 IA: {fonte} &nbsp;|&nbsp; 🏗️ SpaceIntel RPA",
              "</sub>"]

    path.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"[green]✓ Relatório Markdown gerado: {path}[/green]")
    return path
